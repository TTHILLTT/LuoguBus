import requests
import time
import sys
import json
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def fetch_luogu_submissions(luogu_uid, client_id, count=50):
    """è·å–æ´›è°·æäº¤è®°å½•ï¼Œç¡®ä¿è·å–æœ€æ–°è®°å½•å¹¶æŒ‰æ—¶é—´å‡åºæ’åˆ—"""
    url = "https://www.luogu.com.cn/record/list"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Referer": f"https://www.luogu.com.cn/user/{luogu_uid}",
        "Cookie": f"__client_id={client_id}; _uid={luogu_uid}",
        "X-Luogu-Type": "content-only",
        "X-Requested-With": "XMLHttpRequest"
    }

    all_records = []
    page = 1
    per_page = 20  # æ´›è°·æ¯é¡µå›ºå®šè¿”å›20æ¡è®°å½•
    total_needed = count

    try:
        # è®¡ç®—éœ€è¦æŠ“å–çš„é¡µæ•°ï¼ˆå‘ä¸Šå–æ•´ï¼‰
        pages_needed = (total_needed + per_page - 1) // per_page

        # ä»ç¬¬ä¸€é¡µå¼€å§‹è·å–æœ€æ–°è®°å½•
        while len(all_records) < total_needed and page <= pages_needed:
            params = {
                "user": luogu_uid,
                "page": page,
                "_contentOnly": 1
            }

            response = requests.get(url, headers=headers, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()

            if data.get('code', 0) != 200:
                error_msg = data.get('currentData', {}).get('errorMessage', 'æœªçŸ¥é”™è¯¯')
                print(f"ç¬¬ {page} é¡µAPIé”™è¯¯: {error_msg}")
                break

            records = data['currentData']['records']['result']
            if not records:
                break  # æ²¡æœ‰æ›´å¤šè®°å½•äº†

            # å°†æœ¬é¡µè®°å½•æŒ‰æ—¶é—´å‡åºæ’åˆ—åæ·»åŠ åˆ°æ€»åˆ—è¡¨
            records.sort(key=lambda x: x['submitTime'])
            all_records.extend(records)

            # æ˜¾ç¤ºè¿›åº¦
            print(f"å·²è·å–ç¬¬ {page} é¡µï¼Œå…± {len(records)} æ¡è®°å½•")

            # å¦‚æœå·²ç»è·å–åˆ°è¶³å¤Ÿæ•°é‡çš„è®°å½•
            if len(all_records) >= total_needed:
                break

            page += 1  # è·å–ä¸‹ä¸€é¡µ
            time.sleep(0.5)  # é¿å…è¯·æ±‚è¿‡äºé¢‘ç¹

        # æŒ‰æ—¶é—´å‡åºæ’åºï¼ˆè¶Šæ—©çš„è®°å½•è¶Šé å‰ï¼‰
        all_records.sort(key=lambda x: x['submitTime'])

        # ç¡®ä¿ä¸è¶…è¿‡è¯·æ±‚çš„æ•°é‡ï¼ˆå–æœ€æ–°çš„countæ¡ï¼‰
        return all_records[-min(total_needed, len(all_records)):]
    except Exception as e:
        print(f"è·å–æ•°æ®å¤±è´¥: {str(e)}")
        return []


def create_excel(records, filename):
    """åˆ›å»ºç¾è§‚çš„Excelæ–‡ä»¶"""
    wb = Workbook()
    ws = wb.active
    ws.title = "åˆ·é¢˜è®°å½•"

    # è®¾ç½®åˆ—å®½
    column_widths = {
        'A': 20, 'B': 15, 'C': 40,
        'D': 10, 'E': 15, 'F': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # è¡¨å¤´æ ·å¼
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # å†™å…¥è¡¨å¤´
    headers = [
        "æäº¤æ—¥æœŸ", "é¢˜å·", "é¢˜ç›®åç§°", "çŠ¶æ€", "è¿è¡Œæ—¶é—´", "å†…å­˜å ç”¨"
    ]
    ws.append(headers)

    # åº”ç”¨è¡¨å¤´æ ·å¼
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # æ•°æ®æ ·å¼
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    status_colors = {
        "AC": "00B050",  # ç»¿è‰²
        "WA": "FF0000",  # çº¢è‰²
        "TLE": "FFC000",  # æ©™è‰²
        "MLE": "7030A0",  # ç´«è‰²
        "RE": "FF0000",  # çº¢è‰²
        "CE": "000000",  # é»‘è‰²
        "Half AC": "00CED1",  # é’è‰²
        "UKE": "000000"  # é»‘è‰²
    }

    # å†™å…¥æ•°æ®
    for record in records:
        submit_time = datetime.fromtimestamp(record['submitTime'])
        problem = record.get('problem', {})

        # çŠ¶æ€å¤„ç†ï¼ˆåŒ…å«UKEçŠ¶æ€ï¼‰
        status = {
            12: "AC", 7: "WA", 4: "TLE",
            5: "MLE", 6: "RE", 2: "CE",
            14: "Half AC", 11: "UKE"
        }.get(record['status'], f"æœªçŸ¥({record['status']})")

        row = [
            submit_time.strftime("%Y-%m-%d %H:%M"),
            problem.get('pid', 'æœªçŸ¥'),
            problem.get('title', 'æœªçŸ¥'),
            status,
            f"{record.get('time', 0)}ms",
            f"{record.get('memory', 0)}KB",
        ]
        ws.append(row)

        # åº”ç”¨æ•°æ®æ ·å¼
        last_row = ws.max_row
        for col in range(1, 7):  # å…±6åˆ—
            cell = ws.cell(row=last_row, column=col)
            cell.alignment = data_alignment
            cell.border = thin_border

            # çŠ¶æ€åˆ—ç€è‰²ï¼ˆç¬¬4åˆ—ï¼‰
            if col == 4 and status in status_colors:
                cell.fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type="solid"
                )
                # æ·±è‰²èƒŒæ™¯ä½¿ç”¨ç™½è‰²å­—ä½“
                cell.font = Font(color="FFFFFF" if status in ["WA", "RE", "CE", "UKE"] else "000000")

    # å†»ç»“é¦–è¡Œ
    ws.freeze_panes = "A2"

    # ä¿å­˜æ–‡ä»¶
    wb.save(filename)
    print(f"âœ“ å·²ç”ŸæˆExcelæ–‡ä»¶: {filename}")


def create_csv(records, filename):
    """åˆ›å»ºCSVæ–‡ä»¶"""
    if not records:
        return

    # å‡†å¤‡è¡¨å¤´
    fieldnames = [
        "submit_time", "problem_id", "problem_name", "status", "run_time", "memory_usage"
    ]

    # å‡†å¤‡æ•°æ®
    data = []
    for record in records:
        submit_time = datetime.fromtimestamp(record['submitTime'])
        problem = record.get('problem', {})

        status = {
            12: "AC", 7: "WA", 4: "TLE",
            5: "MLE", 6: "RE", 2: "CE",
            14: "Half AC", 11: "UKE"
        }.get(record['status'], f"Unknown({record['status']})")

        data.append({
            "submit_time": submit_time.strftime("%Y-%m-%d %H:%M:%S"),
            "problem_id": problem.get('pid', 'Unknown'),
            "problem_name": problem.get('title', 'Unknown'),
            "status": status,
            "run_time": record.get('time', 0),
            "memory_usage": record.get('memory', 0)
        })

    # å†™å…¥CSV
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

    print(f"âœ“ å·²ç”ŸæˆCSVæ–‡ä»¶: {filename}")


def main():
    banner = f"""
                â–ˆâ–ˆâ•—     â–ˆâ–ˆâ•—   â–ˆâ–ˆâ•— â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—  â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•— â–ˆâ–ˆâ•—   â–ˆâ–ˆâ•—
                â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â•â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ•”â•â•â•â•â• â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘
                â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘  â–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘
                â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘  
                â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â•šâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â•šâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â•šâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â•šâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•
                â•šâ•â•â•â•â•â•â• â•šâ•â•â•â•â•â•  â•šâ•â•â•â•â•â•  â•šâ•â•â•â•â•â•  â•šâ•â•â•â•â•â•
                â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•— â–ˆâ–ˆâ•—   â–ˆâ–ˆâ•—â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—
                â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â•â•â•â•â•
                â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—
                â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘â•šâ•â•â•â•â–ˆâ–ˆâ•‘
                â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â•šâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•‘
                â•šâ•â•â•â•â•â•  â•šâ•â•â•â•â•â• â•šâ•â•â•â•â•â•â•
            """
    print(banner)
    print("æ´›è°·åšé¢˜æ—¥è®°ç”Ÿæˆå™¨ v3.8")
    print("=" * 60)
    print("ä¼˜åŒ–è®°å½•è·å–é€»è¾‘ï¼šç¡®ä¿è·å–æœ€æ–°æäº¤å¹¶æŒ‰æ—¶é—´å‡åºæ’åˆ—")
    print("æ”¯æŒå¤šé¡µæŠ“å–ï¼Œæœ€å¤šå¯è·å–1000æ¡æäº¤è®°å½•ï¼ˆé¿å…é¢‘ç¹è¯·æ±‚è¢«Banï¼‰")
    print("=" * 60)

    # è·å–å¿…è¦ä¿¡æ¯
    client_id = input("è¯·è¾“å…¥__client_idçš„å€¼: ").strip()
    luogu_uid = input("è¯·è¾“å…¥_uidçš„å€¼: ").strip()

    # è¯¢é—®è®°å½•æ•°é‡ï¼ˆæœ€å¤§1000ï¼‰
    while True:
        try:
            count_input = input("è¯·è¾“å…¥è¦è·å–çš„è®°å½•æ•°é‡ (1-1000, é»˜è®¤50): ").strip()
            if not count_input:
                count = 50
                break

            count = int(count_input)
            if 1 <= count <= 1000:
                break
            else:
                print("è¯·è¾“å…¥1åˆ°1000ä¹‹é—´çš„æ•´æ•°ï¼")
        except ValueError:
            print("è¯·è¾“å…¥æœ‰æ•ˆçš„æ•´æ•°ï¼")

    if not client_id or not luogu_uid:
        print("é”™è¯¯: å¿…é¡»æä¾›Cookieä¿¡æ¯")
        sys.exit(1)

    # è·å–æäº¤è®°å½•
    print(f"\næ­£åœ¨è·å–ç”¨æˆ· {luogu_uid} çš„æœ€æ–° {count} æ¡æäº¤è®°å½•...")
    records = fetch_luogu_submissions(luogu_uid, client_id, count)

    if not records:
        print("è·å–æäº¤è®°å½•å¤±è´¥ï¼Œè¯·æ£€æŸ¥ï¼š")
        print("1. Cookieä¿¡æ¯æ˜¯å¦æ­£ç¡®ï¼ˆéœ€åŒ…å«__client_idå’Œ_uidï¼‰")
        print("2. è´¦å·æ˜¯å¦æœ‰å…¬å¼€æäº¤è®°å½•")
        sys.exit(1)

    actual_count = len(records)
    if actual_count < count:
        print(f"âš ï¸ æ³¨æ„: åªè·å–åˆ° {actual_count} æ¡è®°å½•ï¼ˆè¯·æ±‚æ•°é‡: {count}ï¼‰")
    else:
        print(f"âœ… æˆåŠŸè·å– {actual_count} æ¡æäº¤è®°å½•")
        # æ˜¾ç¤ºæ—¶é—´èŒƒå›´
        first_submit = datetime.fromtimestamp(records[0]['submitTime']).strftime("%Y-%m-%d %H:%M")
        last_submit = datetime.fromtimestamp(records[-1]['submitTime']).strftime("%Y-%m-%d %H:%M")
        print(f"ğŸ“… æ—¶é—´èŒƒå›´: {first_submit} è‡³ {last_submit}")

    # ç”Ÿæˆæ–‡ä»¶å
    timestamp = time.strftime('%Y%m%d_%H%M%S')
    base_filename = f"Luogu_Diary_{luogu_uid}_{timestamp}"

    # ç”Ÿæˆä¸¤ç§æ ¼å¼æ–‡ä»¶
    create_excel(records, f"{base_filename}.xlsx")
    create_csv(records, f"{base_filename}.csv")

    # ä½¿ç”¨æç¤º
    print("\nä½¿ç”¨è¯´æ˜:")
    print(f"1. Excelæ–‡ä»¶ ({base_filename}.xlsx):")
    print("   - è¡¨æ ¼æŒ‰æäº¤æ—¶é—´å‡åºæ’åˆ—ï¼ˆæœ€æ—©çš„åœ¨æœ€ä¸Šé¢ï¼‰")
    print("   - ç¾è§‚çš„æ ¼å¼åŒ–è¡¨æ ¼ï¼ŒçŠ¶æ€è‡ªåŠ¨ç€è‰²")
    print(f"2. CSVæ–‡ä»¶ ({base_filename}.csv):")
    print("   - çº¯æ–‡æœ¬æ ¼å¼ï¼Œé€‚åˆç¨‹åºå¤„ç†")
    print(f"3. åŒ…å« {actual_count} æ¡è®°å½•ï¼Œæ—¶é—´ä» {first_submit} åˆ° {last_submit}")
    print("\næç¤ºï¼šé¿å…é¢‘ç¹è¯·æ±‚å¤§é‡æ•°æ®ï¼Œä»¥é˜²è¢«æ´›è°·å°ç¦ï¼")


if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("é”™è¯¯: éœ€è¦å®‰è£…openpyxlåº“ï¼Œè¯·æ‰§è¡Œ: pip install openpyxl")
        sys.exit(1)
    try:
        import requests
    except ImportError:
        print("é”™è¯¯: éœ€è¦å®‰è£…requestsåº“ï¼Œè¯·æ‰§è¡Œ: pip install requests")
        sys.exit(1)

    main()
