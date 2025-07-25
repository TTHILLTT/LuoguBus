import os
import re
import sqlite3
import json
import csv
import time
import sys
import requests
import winreg
import shutil
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.backends import default_backend
import base64


# 浏览器Cookie提取工具
class BrowserCookieExtractor:
    @staticmethod
    def get_chrome_cookies(domain):
        """从Chrome浏览器提取指定域名的Cookie"""
        try:
            # 获取Chrome用户数据目录
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                r"Software\Google\Chrome\NativeMessagingHosts\com.google.chrome.example.echo") as key:
                chrome_path = winreg.QueryValue(key, None)
                user_data_dir = os.path.join(os.path.dirname(chrome_path), "User Data")

            # 查找最新的Cookie文件
            cookie_files = []
            for root, _, files in os.walk(user_data_dir):
                if "Cookies" in files:
                    cookie_files.append(os.path.join(root, "Cookies"))

            if not cookie_files:
                return {}

            # 使用最新的Cookie文件
            cookie_db = max(cookie_files, key=os.path.getmtime)

            # 复制Cookie文件到临时目录（避免锁定问题）
            temp_dir = tempfile.mkdtemp()
            temp_cookie_file = os.path.join(temp_dir, "Cookies")
            shutil.copyfile(cookie_db, temp_cookie_file)

            # 连接到SQLite数据库
            conn = sqlite3.connect(temp_cookie_file)
            cursor = conn.cursor()

            # 查询Cookie
            cursor.execute("""
                SELECT host_key, name, value, encrypted_value 
                FROM cookies 
                WHERE host_key LIKE ? OR host_key LIKE ?
            """, (f"%{domain}", f".%{domain}"))

            # 获取加密密钥
            key = BrowserCookieExtractor.get_chrome_key()

            cookies = {}
            for host_key, name, value, encrypted_value in cursor.fetchall():
                if encrypted_value:
                    decrypted_value = BrowserCookieExtractor.decrypt_chrome_value(encrypted_value, key)
                    if decrypted_value:
                        cookies[name] = decrypted_value
                elif value:
                    cookies[name] = value

            conn.close()
            shutil.rmtree(temp_dir)
            return cookies
        except Exception as e:
            print(f"提取Cookie失败: {str(e)}")
            return {}

    @staticmethod
    def get_chrome_key():
        """获取Chrome加密密钥"""
        try:
            # 获取本地状态文件路径
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                r"Software\Google\Chrome\NativeMessagingHosts\com.google.chrome.example.echo") as key:
                chrome_path = winreg.QueryValue(key, None)
                local_state_file = os.path.join(os.path.dirname(chrome_path), "User Data", "Local State")

            # 读取加密密钥
            with open(local_state_file, "r", encoding="utf-8") as f:
                local_state = json.load(f)

            encrypted_key = base64.b64decode(local_state["os_crypt"]["encrypted_key"])
            encrypted_key = encrypted_key[5:]  # 移除DPAPI前缀

            # 使用Windows DPAPI解密密钥
            import ctypes
            import ctypes.wintypes

            class DATA_BLOB(ctypes.Structure):
                _fields_ = [("cbData", ctypes.wintypes.DWORD),
                            ("pbData", ctypes.POINTER(ctypes.c_char))]

            blob = DATA_BLOB()
            blob.cbData = len(encrypted_key)
            blob.pbData = ctypes.cast(ctypes.create_string_buffer(encrypted_key), ctypes.POINTER(ctypes.c_char))

            ctypes.windll.crypt32.CryptUnprotectData.restype = ctypes.c_int
            ctypes.windll.crypt32.CryptUnprotectData.argtypes = [
                ctypes.POINTER(DATA_BLOB), ctypes.c_wchar_p,
                ctypes.POINTER(DATA_BLOB), ctypes.c_void_p,
                ctypes.c_void_p, ctypes.c_int, ctypes.POINTER(DATA_BLOB)
            ]

            out_blob = DATA_BLOB()
            result = ctypes.windll.crypt32.CryptUnprotectData(
                ctypes.byref(blob), None, None, None, None, 0, ctypes.byref(out_blob)
            )

            if not result:
                raise Exception("DPAPI解密失败")

            key = ctypes.string_at(out_blob.pbData, out_blob.cbData)
            return key
        except Exception as e:
            print(f"获取Chrome密钥失败: {str(e)}")
            return None

    @staticmethod
    def decrypt_chrome_value(encrypted_value, key):
        """解密Chrome Cookie值"""
        try:
            # 提取初始向量和加密数据
            iv = encrypted_value[3:15]
            encrypted_value = encrypted_value[15:]

            # 创建解密器
            cipher = Cipher(algorithms.AES(key), modes.GCM(iv), backend=default_backend())
            decryptor = cipher.decryptor()

            # 解密数据
            decrypted_value = decryptor.update(encrypted_value) + decryptor.finalize()
            return decrypted_value.decode("utf-8")
        except Exception as e:
            print(f"解密失败: {str(e)}")
            return None


# 洛谷数据获取工具
class LuoguDataFetcher:
    @staticmethod
    def fetch_submissions(luogu_uid, client_id, count=50):
        """获取洛谷提交记录"""
        url = "https://www.luogu.com.cn/record/list"

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
            "Referer": f"https://www.luogu.com.cn/user/{luogu_uid}",
            "Cookie": f"__client_id={client_id}; _uid={luogu_uid}",
            "X-Luogu-Type": "content-only",
            "X-Requested-With": "XMLHttpRequest"
        }

        params = {
            "user": luogu_uid,
            "page": 1,
            "_contentOnly": 1
        }

        try:
            response = requests.get(url, headers=headers, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()

            if data.get('code', 0) != 200:
                print(f"API错误: {data.get('currentData', {}).get('errorMessage', '未知错误')}")
                return []

            records = data['currentData']['records']['result']
            records.sort(key=lambda x: x['submitTime'])  # 按时间升序排序
            return records[:min(count, len(records))]
        except Exception as e:
            print(f"获取数据失败: {str(e)}")
            return []


# 文件导出工具
class DiaryExporter:
    @staticmethod
    def create_excel(records, filename):
        """创建美观的Excel文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "刷题记录"

        # 设置列宽
        column_widths = {
            'A': 25, 'B': 25, 'C': 25,
        'D': 25, 'E': 25, 'F': 25
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # 写入表头
        headers = [
            "提交日期", "题号", "题目名称", "状态","运行时间", "内存占用"
        ]
        ws.append(headers)

        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # 数据样式
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        status_colors = {
            "AC": "00B050",  # 绿色
            "WA": "FF0000",  # 红色
            "TLE": "FFC000",  # 橙色
            "MLE": "7030A0",  # 紫色
            "RE": "FF0000",  # 红色
            "CE": "000000"  # 黑色
        }

        # 写入数据
        for record in records:
            submit_time = datetime.fromtimestamp(record['submitTime'])
            problem = record.get('problem', {})

            # 状态和难度处理
            status = {
                12: "AC", 7: "WA", 4: "TLE",
                5: "MLE", 6: "RE", 2: "CE", 14: "Half AC"
            }.get(record['status'], f"未知({record['status']})")

            row = [
                submit_time.strftime("%Y-%m-%d %H:%M"),
                problem.get('pid', '未知'),
                problem.get('title', '未知'),
                status,
                f"{record.get('time', 0)}ms",
                f"{record.get('memory', 0)}KB"
            ]
            ws.append(row)

            # 应用数据样式
            last_row = ws.max_row
            for col in range(1, 11):
                cell = ws.cell(row=last_row, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border

                # 状态列着色
                if col == 5 and status in status_colors:
                    cell.fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status],
                                            fill_type="solid")
                    cell.font = Font(color="FFFFFF" if status in ["WA", "RE", "CE"] else "000000")

        # 冻结首行
        ws.freeze_panes = "A2"

        # 保存文件
        wb.save(filename)
        print(f"✓ 已生成Excel文件: {filename}")

    @staticmethod
    def create_csv(records, filename):
        """创建CSV文件"""
        if not records:
            return

        # 准备表头
        fieldnames = [
            "submit_time", "problem_id", "problem_name","status", "run_time", "memory_usage"
        ]

        # 准备数据
        data = []
        for record in records:
            submit_time = datetime.fromtimestamp(record['submitTime'])
            problem = record.get('problem', {})

            status = {
                12: "AC", 7: "WA", 4: "TLE",
                5: "MLE", 6: "RE", 2: "CE", 14: "Half AC"
            }.get(record['status'], f"Unknown({record['status']})")

            data.append({
                "submit_time": submit_time.strftime("%Y-%m-%d %H:%M:%S"),
                "problem_id": problem.get('pid', 'Unknown'),
                "problem_name": problem.get('title', 'Unknown'),
                "status": status,
                "run_time": record.get('time', 0),
                "memory_usage": record.get('memory', 0)
            })

            data.append({
                "submit_time": submit_time.strftime("%Y-%m-%d %H:%M:%S"),
                "problem_id": problem.get('pid', 'Unknown'),
                "problem_name": problem.get('title', 'Unknown'),
                "status": status,
                "run_time": record.get('time', 0),
                "memory_usage": record.get('memory', 0)
            })

        # 写入CSV
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

        print(f"✓ 已生成CSV文件: {filename}")


# 主程序
class LuoguDiaryGenerator:
    def __init__(self):
        self.version = "3.1Cookie"
        self.banner = f"""
            ██╗     ██╗   ██╗ ██████╗  ██████╗ ██╗   ██╗
            ██║     ██║   ██║██╔════╝ ██╔═══██╗██║   ██║
            ██║     ██║   ██║██║  ███╗██║   ██║██║   ██║
            ██║     ██║   ██║██║   ██║██║   ██║██║   ██║  
            ███████╗╚██████╔╝╚██████╔╝╚██████╔╝╚██████╔╝
            ╚══════╝ ╚═════╝  ╚═════╝  ╚═════╝  ╚═════╝
            ██████╗ ██╗   ██╗███████╗
            ██╔══██╗██║   ██║██╔════╝
            ██████╔╝██║   ██║███████╗
            ██╔══██╗██║   ██║╚════██║
            ██████╔╝╚██████╔╝███████║
            ╚═════╝  ╚═════╝ ╚══════╝
        洛谷刷题日记生成器 v{self.version}
        """

    def run(self):
        print(self.banner)
        print("=" * 80)
        print("功能特点:")
        print("1. 🪄 自动从浏览器提取Cookie（无需手动查找）")
        print("2. 📊 同时生成Excel和CSV格式的刷题日记")
        print("3. 🕒 提交记录按时间升序排列（越早越靠上）")
        print("4. 🎨 Excel自动美化（状态着色、格式优化）")
        print("=" * 80)

        # 自动获取Cookie
        print("\n正在尝试从浏览器提取Cookie...")
        cookies = BrowserCookieExtractor.get_chrome_cookies("luogu.com.cn")

        client_id = cookies.get("__client_id", "")
        luogu_uid = cookies.get("_uid", "")

        if not client_id or not luogu_uid:
            print("\n⚠️ 自动提取失败，请手动输入Cookie信息")
            print("=" * 60)
            print("如何获取Cookie信息:")
            print("1. 登录洛谷网站")
            print("2. 按F12打开开发者工具")
            print("3. 转到Application/存储 > Cookies")
            print("4. 找到__client_id和_uid的值")
            print("=" * 60)

            client_id = input("请输入__client_id的值: ").strip()
            luogu_uid = input("请输入_uid的值: ").strip()
        else:
            print("✅ 成功提取Cookie信息!")

        if not client_id or not luogu_uid:
            print("错误: 必须提供Cookie信息")
            sys.exit(1)

        # 获取提交记录
        print(f"\n🔍 正在获取用户 {luogu_uid} 的提交记录...")
        records = LuoguDataFetcher.fetch_submissions(luogu_uid, client_id, 100)

        if not records:
            print("获取提交记录失败，请检查：")
            print("1. Cookie信息是否正确（需包含__client_id和_uid）")
            print("2. 账号是否有公开提交记录")
            sys.exit(1)

        # 生成文件
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        base_filename = f"洛谷刷题日记_{luogu_uid}_{timestamp}"

        print("\n🛠️ 正在生成文件...")
        DiaryExporter.create_excel(records, f"{base_filename}.xlsx")
        DiaryExporter.create_csv(records, f"{base_filename}.csv")

        # 使用提示
        print("\n" + "=" * 60)
        print("🎉 生成成功！使用说明:")
        print(f"1. Excel文件 ({base_filename}.xlsx):")
        print("   - 美观的格式化表格，适合直接查看和编辑")
        print("   - 状态自动着色（AC绿色/WA红色等）")
        print("   - 请补充'解题用时'和'解题笔记/反思'列")
        print(f"2. CSV文件 ({base_filename}.csv):")
        print("   - 纯文本格式，适合程序处理或导入数据库")
        print("3. 两个文件内容相同，按提交时间升序排列")
        print("=" * 60)
        print("提示：每天训练后运行此程序，持续更新刷题日记！")


if __name__ == "__main__":
    try:
        # 检查必要的库
        import requests
        import openpyxl
        import cryptography

        # 运行生成器
        generator = LuoguDiaryGenerator()
        generator.run()
    except ImportError as e:
        print("错误: 缺少必要的依赖库")
        print("请执行以下命令安装依赖:")
        print("pip install requests openpyxl cryptography pywin32")
        sys.exit(1)
    except Exception as e:
        print(f"程序运行出错: {str(e)}")
        sys.exit(1)